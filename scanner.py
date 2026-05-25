from __future__ import annotations

from pathlib import Path

import pandas as pd
import yfinance as yf
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("reports")
FOCUS_SYMBOLS = [
    "NVDA", "AMD", "AVGO", "TSM", "MU", "MRVL", "PLTR", "ANET", "VRT", "CEG", "CRWD", "MSFT", "META", "GOOGL",
    "AAPL", "AMZN", "TSLA", "COST", "NFLX", "QCOM", "AMAT", "LRCX", "KLAC", "INTC", "SMH", "QQQ", "VOO",
]

SECTOR_ZH = {
    "Technology": "科技",
    "Communication Services": "通信服务",
    "Consumer Cyclical": "可选消费",
    "Consumer Defensive": "必需消费",
    "Financial Services": "金融服务",
    "Healthcare": "医疗保健",
    "Industrials": "工业",
    "Real Estate": "房地产",
    "Energy": "能源",
    "Utilities": "公用事业",
    "Basic Materials": "基础材料",
}

NAME_ZH = {
    "NVDA": "英伟达", "AMD": "超威半导体", "AVGO": "博通", "TSM": "台积电", "MU": "美光科技", "MRVL": "迈威尔科技",
    "PLTR": "Palantir", "ANET": "Arista网络", "VRT": "维谛技术", "CEG": "Constellation Energy", "CRWD": "CrowdStrike",
    "MSFT": "微软", "META": "Meta平台", "GOOGL": "谷歌A", "AAPL": "苹果", "AMZN": "亚马逊", "TSLA": "特斯拉",
    "COST": "好市多", "NFLX": "奈飞", "QCOM": "高通", "AMAT": "应用材料", "LRCX": "拉姆研究", "KLAC": "科磊",
    "INTC": "英特尔", "SMH": "半导体ETF", "QQQ": "纳指100ETF", "VOO": "标普500ETF",
}

COLUMN_WIDTHS = {
    "排名": 8,
    "代码": 12,
    "中文名称": 24,
    "最新收盘价": 14,
    "总分": 10,
    "板块": 16,
    "行业": 28,
    "5日涨跌幅": 14,
    "20日涨跌幅": 14,
    "距52周高点回撤": 20,
    "成交量/20日均量": 20,
    "成交量是否放大": 18,
    "财报日期": 14,
    "买入逻辑": 46,
    "风险点": 46,
}


def pct(value: float | None) -> str:
    if value is None or pd.isna(value):
        return "N/A"
    return f"{value:.2%}"


def fetch_symbols() -> list[str]:
    symbols = set(FOCUS_SYMBOLS)
    try:
        tables = pd.read_html("https://en.wikipedia.org/wiki/List_of_S%26P_500_companies")
        symbols.update(tables[0]["Symbol"].astype(str).str.replace(".", "-", regex=False).tolist())
    except Exception as exc:
        print(f"S&P 500 pool fallback only: {exc}")
    try:
        query = yf.EquityQuery("eq", ["region", "us"])
        result = yf.screen(query, size=100, sortField="intradaymarketcap", sortAsc=False)
        symbols.update(q["symbol"] for q in result.get("quotes", []) if q.get("symbol"))
    except Exception as exc:
        print(f"Large-cap pool fallback only: {exc}")
    return sorted(symbols)


def latest_metrics(symbol: str) -> dict | None:
    history = yf.download(symbol, period="18mo", interval="1d", auto_adjust=False, progress=False, threads=False)
    if history.empty or len(history) < 260:
        return None
    close = history["Adj Close"] if "Adj Close" in history.columns else history["Close"]
    volume = history["Volume"]
    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]
    if isinstance(volume, pd.DataFrame):
        volume = volume.iloc[:, 0]
    close = close.dropna()
    volume = volume.dropna()
    price = float(close.iloc[-1])
    ma5 = float(close.rolling(5).mean().iloc[-1])
    ma10 = float(close.rolling(10).mean().iloc[-1])
    ma20 = float(close.rolling(20).mean().iloc[-1])
    ret5 = float(close.iloc[-1] / close.iloc[-6] - 1)
    ret20 = float(close.iloc[-1] / close.iloc[-21] - 1)
    high52 = float(close.tail(252).max())
    drawdown52 = float(close.iloc[-1] / high52 - 1)
    volume_ratio = float(volume.iloc[-1] / volume.tail(20).mean())
    sector = None
    industry = None
    earnings_date = None
    company_name = None
    try:
        ticker = yf.Ticker(symbol)
        info = ticker.get_info()
        sector = info.get("sector")
        industry = info.get("industry")
        company_name = info.get("shortName") or info.get("longName")
        calendar = getattr(ticker, "calendar", None)
        if isinstance(calendar, dict) and calendar.get("Earnings Date"):
            earnings_date = pd.to_datetime(calendar["Earnings Date"][0]).date().isoformat()
    except Exception:
        pass
    trend = sum([price > ma5, price > ma10, price > ma20]) * 5 + (8 if ma5 > ma10 > ma20 else 0) + max(0, min(7, ret20 / 0.12 * 7))
    pullback = 15 if 0.03 <= abs(drawdown52) <= 0.15 else 8
    volume_score = 15 if volume_ratio >= 2 else 12 if volume_ratio >= 1.5 else 9 if volume_ratio >= 1.2 else 6 if volume_ratio >= 1 else 3
    risk = (3 if ret5 > 0.15 else 0) + (3 if ret20 > 0.30 else 0) + (3 if price < ma20 else 0)
    score = max(0, min(100, trend + pullback + volume_score + 10 - risk))
    return {
        "代码": symbol,
        "中文名称": NAME_ZH.get(symbol, company_name or "未收录"),
        "最新收盘价": round(price, 2),
        "总分": round(score, 2),
        "板块": SECTOR_ZH.get(sector, sector or "N/A"),
        "行业": industry or "N/A",
        "5日涨跌幅": pct(ret5),
        "20日涨跌幅": pct(ret20),
        "距52周高点回撤": pct(drawdown52),
        "成交量/20日均量": round(volume_ratio, 2),
        "成交量是否放大": "是" if volume_ratio >= 1.5 else "否",
        "财报日期": earnings_date or "N/A",
        "买入逻辑": "；".join(filter(None, ["价格站上20日均线" if price > ma20 else "", "均线多头排列" if ma5 > ma10 > ma20 else "", "成交量放大" if volume_ratio >= 1.5 else ""])),
        "风险点": "；".join(filter(None, ["5日涨幅过快" if ret5 > 0.15 else "", "20日涨幅过大" if ret20 > 0.30 else "", "价格低于20日均线" if price < ma20 else ""])) or "未见明显规则内风险",
    }


def format_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in writer.book.worksheets:
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 26
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        for column_cells in sheet.columns:
            header = str(column_cells[0].value or "")
            letter = get_column_letter(column_cells[0].column)
            fallback_width = max(len(header) + 4, 14)
            sheet.column_dimensions[letter].width = COLUMN_WIDTHS.get(header, fallback_width)
        for row in sheet.iter_rows(min_row=2):
            sheet.row_dimensions[row[0].row].height = 34
            for cell in row:
                header = str(sheet.cell(row=1, column=cell.column).value or "")
                cell.alignment = Alignment(
                    horizontal="left" if header in {"中文名称", "板块", "行业", "买入逻辑", "风险点"} else "center",
                    vertical="center",
                    wrap_text=header in {"买入逻辑", "风险点"},
                )


def build_reports(rows: list[dict]) -> tuple[Path, Path]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "history").mkdir(parents=True, exist_ok=True)
    today = pd.Timestamp.today().date().isoformat()
    df = pd.DataFrame(rows).sort_values("总分", ascending=False).reset_index(drop=True)
    df.insert(0, "排名", range(1, len(df) + 1))
    top = df.head(10)
    excel_path = OUTPUT_DIR / f"daily_scanner_{today}.xlsx"
    md_path = OUTPUT_DIR / f"daily_scanner_{today}.md"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        top.to_excel(writer, sheet_name="今日关注前十", index=False)
        df.to_excel(writer, sheet_name="全部排名", index=False)
        format_workbook(writer)
    lines = [f"# 美股机会雷达日报 - {today}", "", "## 今日最值得关注", ""]
    for _, row in top.iterrows():
        lines += [
            f"### {int(row['排名'])}. {row['代码']} {row['中文名称']} - {row['总分']:.1f}分",
            "",
            f"- 价格：${row['最新收盘价']}",
            f"- 板块/行业：{row['板块']} / {row['行业']}",
            f"- 5日涨跌幅：{row['5日涨跌幅']}",
            f"- 20日涨跌幅：{row['20日涨跌幅']}",
            f"- 距52周高点回撤：{row['距52周高点回撤']}",
            f"- 成交量：{row['成交量是否放大']}，为20日均量 {row['成交量/20日均量']} 倍",
            f"- 财报日期：{row['财报日期']}",
            f"- 买入逻辑：{row['买入逻辑'] or '技术面中性，需人工复盘确认'}",
            f"- 风险点：{row['风险点']}",
            "",
        ]
    md_path.write_text("\n".join(lines), encoding="utf-8")
    df.assign(scan_date=today).to_csv(OUTPUT_DIR / "history" / "daily_scanner_history.csv", index=False, encoding="utf-8-sig")
    return excel_path, md_path


def main() -> None:
    rows = []
    for symbol in fetch_symbols():
        try:
            row = latest_metrics(symbol)
            if row:
                rows.append(row)
        except Exception as exc:
            print(f"Skip {symbol}: {exc}")
    if not rows:
        raise RuntimeError("No valid stock data downloaded.")
    excel_path, md_path = build_reports(rows)
    print(f"Excel report: {excel_path}")
    print(f"Markdown report: {md_path}")


if __name__ == "__main__":
    main()
